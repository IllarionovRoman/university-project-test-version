from datetime import datetime

from django.db.models import Q
from django.shortcuts import render
from django.http import JsonResponse
from django.utils import timezone
from rest_framework import generics
from .models import Event
from .serializers import EventSerializer
from django.utils import timezone
from .products import products
from django.contrib.auth.models import User
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from .models import Product
from .serializers import *
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth.hashers import make_password
from rest_framework import status
from django.http import HttpResponse
from django.views.generic import View
from django_excel import make_response_from_query_sets
from django.http import HttpResponse
from rest_framework.views import APIView
from openpyxl import Workbook
from .serializers import StudentSerializer


class ExportStudentsToExcel(APIView):
    def get(self, request):
        students = Students.objects.all()
        serializer = StudentSerializer(students, many=True)

        # Создание новой книги Excel и листа
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        # Заполнение заголовков таблицы
        headers = ['ID', 'Имя', 'Фамилия', 'Возраст']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Заполнение данных таблицы
        for row_num, student_data in enumerate(serializer.data, 2):
            ws.cell(row=row_num, column=1, value=student_data['_id'])
            ws.cell(row=row_num, column=2, value=student_data['first_name'])
            ws.cell(row=row_num, column=3, value=student_data['last_name'])
            ws.cell(row=row_num, column=4, value=student_data['age'])
            ...

        # Сохранение файла Excel в HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=my_students.xlsx'
        wb.save(response)

        return response

class ExportInventoryToExcel(APIView):
    def get(self, request):
        inventorys = Inventory.objects.all()
        serializer = InventorySerializer(inventorys, many=True)

        # Создание новой книги Excel и листа
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventorys"

        # Заполнение заголовков таблицы
        headers = ['ID', 'название', 'дата', 'описание', 'ид секции', 'количество']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Заполнение данных таблицы
        for row_num, inventory_data in enumerate(serializer.data, 2):
            ws.cell(row=row_num, column=1, value=inventory_data['id'])
            ws.cell(row=row_num, column=2, value=inventory_data['name'])
            ws.cell(row=row_num, column=3, value=inventory_data['date'])
            ws.cell(row=row_num, column=4, value=inventory_data['description'])
            ws.cell(row=row_num, column=5, value=inventory_data['sections'])
            ws.cell(row=row_num, column=6, value=inventory_data['quantity'])
            ...

        # Сохранение файла Excel в HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=my_inventory.xlsx'
        wb.save(response)

        return response
class EventList(generics.ListAPIView):
    serializer_class = EventSerializer

    def get_queryset(self):
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        # Проверка наличия параметров start_date и end_date в запросе
        if start_date and end_date:
            # Преобразование строковых значений в объекты datetime
            start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d'))

            # Получение данных Events, отфильтрованных по дате
            queryset = Event.objects.filter(date__range=(start_datetime, end_datetime))
        else:
            # Если параметры start_date и end_date отсутствуют, возвращаем все данные Events
            queryset = Event.objects.all()

        return queryset

class PodEventList(generics.ListAPIView):
    serializer_class = PodEventSerializer

    def get_queryset(self):
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        # Проверка наличия параметров start_date и end_date в запросе
        if start_date and end_date:
            # Преобразование строковых значений в объекты datetime
            start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d'))

            # Получение данных Events, отфильтрованных по дате
            queryset = PodEvent.objects.filter(date__range=(start_datetime, end_datetime))
        else:
            # Если параметры start_date и end_date отсутствуют, возвращаем все данные Events
            queryset = PodEvent.objects.all()

        return queryset

class InventoryList(generics.ListAPIView):
    serializer_class = InventorySerializer

    def get_queryset(self):
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        # Проверка наличия параметров start_date и end_date в запросе
        if start_date and end_date:
            # Преобразование строковых значений в объекты datetime
            start_datetime = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end_datetime = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d'))

            # Получение данных Events, отфильтрованных по дате
            queryset = Inventory.objects.filter(date__range=(start_datetime, end_datetime))
        else:
            # Если параметры start_date и end_date отсутствуют, возвращаем все данные Events
            queryset = Inventory.objects.all()

        return queryset
class MyTokenObtainPairSerializer(TokenObtainPairSerializer):

    def validate(self, attrs):
        data = super().validate(attrs)
        serializer = UserSerializerWithToken(self.user).data
        for k, v in serializer.items():
            data[k] = v
        return data


class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer


@api_view(['POST'])
def registerUser(request):
    data = request.data
    try:
        user = User.objects.create(
            first_name=data['name'],
            username=data['email'],
            email=data['email'],
            password=make_password(data['password'])
        )
        serializer = UserSerializerWithToken(user, many=False)
        return Response(serializer.data)
    except:
        message = {'a': 'zanyato'}
        return Response(message, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateUserProfile(request):
    user = request.user
    serializer = UserSerializerWithToken(user, many=False)
    data = request.data
    user.first_name = data['name']
    user.username = data['email']
    user.email = data['email']
    if data['password'] != '':
        user.password = make_password(data['password'])

    user.save()

    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getUserProfile(request):
    user = request.user
    serializer = UserSerializer(user, many=False)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def getUsers(request):
    query = request.query_params.get('us')
    if query == None:
        query = ''
    users = User.objects.filter(username__icontains=query)
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def getProducts(request):

    query = request.query_params.get('keyword')


    if query == None:
        query = ''
    products = Product.objects.filter(name__icontains=query)
    serializer = ProductSerializer(products, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def getProduct(request, pk):
    product = Product.objects.get(_id=pk)
    serializer = ProductSerializer(product, many=False)
    return Response(serializer.data)


@api_view(['GET'])
def getUserById(request, pk):
    user = User.objects.get(id=pk)
    serializer = UserSerializer(user, many=False)
    return Response(serializer.data)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def updateUser(request, pk):
    user = User.objects.get(id=pk)
    data = request.data
    user.first_name = data['name']
    user.username = data['email']
    user.email = data['email']
    user.is_staff = data['isAdmin']
    user.save()
    serializer = UserSerializer(user, many=False)

    return Response(serializer.data)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def deleteUser(request, pk):
    userForDeletion = User.objects.get(id=pk)
    userForDeletion.delete()
    return Response('YDALENO')


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def deleteProduct(request, pk):
    product = Product.objects.get(_id=pk)
    product.delete()
    return Response('удалено')


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def createProduct(request):
    user = request.user
    product = Product.objects.create(
        user=user,
        name='pr',
        description='f',
        price='0.00',
        image='',

    )
    serializer = ProductSerializer(product, many=False)
    return Response(serializer.data)


@api_view(['PUT'])
def updateProduct(request, pk):
    data = request.data
    product = Product.objects.get(_id=pk)
    product.name = data['name']
    product.description = data['description']
    product.price = data['price']
    product.image = data['image']
    product.brand = data['brand']
    product.save()
    serializer = ProductSerializer(product, many=False)
    return Response(serializer.data)


@api_view(['POST'])
def uploadImage(request):
    data = request.data
    product_id = data['product_id']
    product = Product.objects.get(_id=product_id)
    product.image = request.FILES.get('image')
    product.save()
    serializer = ProductSerializer(product, many=False)
    return Response(serializer.data)


# -------------------------------------------------------------------------------------------------------------#

@api_view(['GET'])
def getStudents(request):
    query = request.query_params.get('stud')
    if query == None:
        query = ''
    students = Students.objects.filter(first_name__icontains=query)
    serializer = StudentSerializer(students, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def getStudent(request, pk):
    student = Students.objects.get(_id=pk)
    serializer = StudentSerializer(student, many=False)
    return Response(serializer.data)


@api_view(['POST'])
def createStudent(request):
    user = request.user
    trainers = Students.trainer
    student = Students.objects.create(
        first_name='2',
        last_name='2',
        middle_name='3',
        age='4',

    )
    serializer = StudentSerializer(student, many=False)
    return Response(serializer.data)


@api_view(['GET'])
def getStudent(request, pk):
    student = Students.objects.get(_id=pk)
    serializer = StudentSerializer(student, many=False)
    return Response(serializer.data)


@api_view(['PUT'])
def updateStudent(request, pk):
    data = request.data
    student = Students.objects.get(_id=pk)
    student.first_name = data['first_name']
    student.last_name = data['last_name']
    student.middle_name = data['middle_name']
    student.age = data['age']
    student.save()
    serializer = StudentSerializer(student, many=False)
    return Response(serializer.data)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def deleteStudent(request, pk):
    student = Students.objects.get(_id=pk)
    student.delete()
    return Response('удалено')


# -------------------------------------------------------------------------------------------------------------#

@api_view(['GET'])
def getDepartments(request):
    query = request.query_params.get('dep')
    if query == None:
        query = ''
    department = Departments.objects.filter(name__icontains=query)
    serializer = DepartmentSerializer(department, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def createDepartment(request):
    user = request.user
    department = Departments.objects.create(
        name='dep'
    )
    serializer = DepartmentSerializer(department, many=False)
    return Response(serializer.data)


@api_view(['GET'])
def getDepartment(request, pk):
    department = Departments.objects.get(_id=pk)
    serializer = DepartmentSerializer(department, many=False)
    return Response(serializer.data)


@api_view(['PUT'])
def updateDepartment(request, pk):
    data = request.data
    department = Departments.objects.get(_id=pk)
    department.name = data['name']
    department.save()
    serializer = DepartmentSerializer(department, many=False)
    return Response(serializer.data)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def deleteDepartment(request, pk):
    department = Departments.objects.get(_id=pk)
    department.delete()
    return Response('удалено')


# -------------------------------------------------------------------------------------------------------------#

@api_view(['GET'])
def getSections(request):
    query = request.query_params.get('sec')
    if query == None:
        query = ''
    section = Sections.objects.filter(name__icontains=query)
    serializer = SectionSerializer(section, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def createSection(request):
    user = request.user
    section = Sections.objects.create(
        name='sec',
    )
    serializer = DepartmentSerializer(section, many=False)
    return Response(serializer.data)


@api_view(['GET'])
def getSection(request, pk):
    section = Sections.objects.get(_id=pk)
    serializer = SectionSerializer(section, many=False)
    return Response(serializer.data)


@api_view(['PUT'])
def updateSection(request, pk):
    data = request.data
    section = Sections.objects.get(_id=pk)
    section.name = data['name']
    section.save()
    serializer = DepartmentSerializer(section, many=False)
    return Response(serializer.data)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def deleteSection(request, pk):
    section = Sections.objects.get(_id=pk)
    section.delete()
    return Response('удалено')


# -------------------------------------------------------------------------------------------------------------#

@api_view(['GET'])
def getTrainers(request):
    query = request.query_params.get('trai')
    if query == None:
        query = ''
    trainer = Trainers.objects.filter(first_name__icontains=query)
    serializer = TrainerSerializer(trainer, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def createTrainer(request):
    user = request.user
    trainer = Trainers.objects.create(
        first_name='sec',
        last_name='fa',
        middle_name='fafaf',
    )
    serializer = TrainerSerializer(trainer, many=False)
    return Response(serializer.data)


@api_view(['GET'])
def getTrainer(request, pk):
    trainer = Trainers.objects.get(_id=pk)
    serializer = TrainerSerializer(trainer, many=False)
    return Response(serializer.data)


@api_view(['PUT'])
def updateTrainer(request, pk):
    data = request.data
    trainer = Trainers.objects.get(_id=pk)
    trainer.first_name = data['first_name']
    trainer.last_name = data['last_name']
    trainer.middle_name = data['middle_name']
    trainer.save()
    serializer = TrainerSerializer(trainer, many=False)
    return Response(serializer.data)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def deleteTrainer(request, pk):
    trainer = Trainers.objects.get(_id=pk)
    trainer.delete()
    return Response('удалено')


# -------------------------------------------------------------------------------------------------------------#

@api_view(['GET'])
def getAwards(request):
    query = request.query_params.get('aw')
    if query == None:
        query = ''
    award = Awards.objects.filter(name__icontains=query)
    serializer = AwardSerializer(award, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def createAward(request):
    user = request.user
    award = Awards.objects.create(
        name='sec',
    )
    serializer = AwardSerializer(award, many=False)
    return Response(serializer.data)


@api_view(['GET'])
def getAward(request, pk):
    award = Awards.objects.get(_id=pk)
    serializer = AwardSerializer(award, many=False)
    return Response(serializer.data)


@api_view(['PUT'])
def updateAward(request, pk):
    data = request.data
    award = Awards.objects.get(_id=pk)
    award.name = data['name']
    award.save()
    serializer = AwardSerializer(award, many=False)
    return Response(serializer.data)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def deleteAward(request, pk):
    award = Awards.objects.get(_id=pk)
    award.delete()
    return Response('удалено')




